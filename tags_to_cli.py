###____________________________________________________________________________
###
###	csv	to	cli	cloud				2021'10 botao
###____________________________________________________________________________
###

import sys
import argparse
import openpyxl
from distutils.util import strtobool

###____________________________________________________________________________

def dump_worksheet(ws):
    """ dump excel worksheet """
    print('')
    print("============================================================")
    print("=== worksheet : " + ws.title)
    print("============================================================")
    ### print (ws)
    for i in range(0, ws.max_row):
        print("=== line (%i)" % (1+i))
        for col in ws.iter_cols(1, ws.max_column):
            print(col[i].value, end=";")
        print('<eol>')
    print('<eows>')

###____________________________________________________________________________

if __name__ == "__main__":

    parser = argparse.ArgumentParser(description="""
    generate cloud cli commands from tags in xlsx file. 
    """)
    ###________________________________________________________________________

    parser.add_argument("--cloud",  default="az", choices=["aws", "az", "gcp", "oci"], help="cloud provider")
    parser.add_argument("--output", default="tags.out", help="output file name")
    parser.add_argument("--xlsx",   default="tags.xlsx", help="excel file name")

    parser.add_argument("--wsname", default="undef", help="worksheet name")

    parser.add_argument("--dbsave", default=False, type=strtobool, help="store in db ? (default: %(default)s)")
    parser.add_argument("--dbname", default="tags.db", help="db file name")
    parser.add_argument("--dbtype", default="sqlite", choices=["mysql", "sqlite"], help="db brand")

    parser.add_argument("--dump",   default=False, type=strtobool, help="dump workbook/worksheet ? (default: %(default)s)")

    args = parser.parse_args()
    ###________________________________________________________________________

    CLOUDNAME  = args.cloud
    OUTPUTFILE = args.output
    XLSXNAME   = args.xlsx

    WSNAME     = args.wsname

    SAVEFLAG   = args.dbsave
    DBNAME     = args.dbname
    DBTYPE     = args.dbtype

    DUMPFLAG   = args.dump

    print('')
    print("=== cloud  : " + CLOUDNAME)
    print("=== output : " + OUTPUTFILE)
    print("=== xlsx   : " + XLSXNAME)

    if SAVEFLAG:
        print("=== dbname : " + DBNAME)
        print("=== dbtype : " + DBTYPE)
    ###________________________________________________________________________

    workbook = openpyxl.load_workbook(XLSXNAME)

    print('')
    print("============================================================")
    ### print(workbook.get_sheet_names())
    print("=== sheet names : %s" % (workbook.sheetnames))
    print("============================================================")

    if DUMPFLAG:
        if WSNAME == "undef":
            for worksheet in workbook.worksheets:
                dump_worksheet(worksheet)
        else:
            worksheet = workbook.get_sheet_by_name(WSNAME)
            dump_worksheet(worksheet)

###____________________________________________________________________________


### 
###           |          _                   |~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~|
###           |\       _/ \_                 |       alexandre  botao       |
###           | \_    /_    \_               |         botao dot org        |
###           \   \__/  \__   \              |       +55-11-98244-UNIX      |
###            \_    \__/  \_  \             |       +55-11-9933-LINUX      |
###              \_   _/     \ |             |  alexandre at botao dot org  |
###                \_/        \|             |      botao at unix  dot net  |
###                            |             |______________________________|
### 


SWNAME="tags_to_cli.py"
SWVERS="1.0.0"
SWDATE="2021-10-24"
SWTIME="00:00:00"
SWDESC="generate cloud cli commands from tags in xlsx file"
SWTAGS="cloud,cli,tags,xlsx,excel,sqlite,mysql,python3"
SWCOPY="GPLv3"
SWAUTH="alexandre botao"
SWMAIL="alexandre at botao dot org"


##  __________________________________________________________________________
## |                                                                          |
## |  This software is free and open-source: you can redistribute it and/or   |
## |  modify it under the terms stated on the GNU General Public License      |
## |  as published by the Free Software Foundation, either version 3 of the   |
## |  License, or (at your option) any later version.                         |
## |                                                                          |
## |   This code is distributed in the hope that it will be useful,           |
## |   but WITHOUT ANY WARRANTY; without even the implied warranty of         |
## |   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.                   |
## |   See the GNU General Public License for more details.                   |
## |                                                                          |
## |   You should have received a copy of the GNU General Public License      |
## |   along with this code.  If not, see <http://www.gnu.org/licenses/>,     |
## |   or write to the Free Software Foundation, Inc.,                        |
## |   59 Temple Place, Suite 330, Boston, MA  02111-1307  USA.               |
## |__________________________________________________________________________|
##


### vi:nu
